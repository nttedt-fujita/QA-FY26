"""CSVデータとExcel元データの整合性テスト

検証項目:
- 各シートの行数が一致するか
- 品名・品番が欠落していないか
- 検査工数(分)の合計が一致するか
- 日付変換が正しいか（単体 + 実データ）
- safe_filenameが正しく動作するか
- カテゴリ列が正しいか
- by-part CSVの行数・工数合計がrawと一致するか
- summary CSVの集計値・ソート順が正しいか
"""

import sys
import os
from pathlib import Path

# プロジェクトルートを取得（tools/tests/incoming_inspection/ から3階層上）
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
# incoming_inspectionモジュールへのパス追加
sys.path.insert(0, str(PROJECT_ROOT / 'tools' / 'incoming_inspection'))

from extract_csv import serial_to_date, safe_filename

import pytest
import openpyxl
import csv
from datetime import datetime

# パス設定（相対パス）
EXCEL_PATH = PROJECT_ROOT / 'docs' / 'excel-orgn' / '受入検査作業集計.xlsx'
OUT_DIR = PROJECT_ROOT / 'sessions' / 'session19' / 'csv-output'
RAW_DIR = OUT_DIR / 'raw'
PART_DIR = OUT_DIR / 'by-part'
SUMMARY_DIR = OUT_DIR / 'summary'

# シート設定（extract_csv.pyのSHEET_CONFIGから必要な列だけ抽出）
SHEET_CONFIG = {
    '検品結果報告（メカ）': {'category': 'メカ', 'data_start': 6, 'hinmei_col': 5, 'hinban_col': 4, 'kosu_col': 13, 'nyuka_col': 1},
    '検品結果報告（エレキ）': {'category': 'エレキ', 'data_start': 6, 'hinmei_col': 4, 'hinban_col': 3, 'kosu_col': 12, 'nyuka_col': 1},
    '検品結果報告（Api）': {'category': 'Api', 'data_start': 6, 'hinmei_col': 4, 'hinban_col': 3, 'kosu_col': 12, 'nyuka_col': 1},
    '受入に時間がかかるもの': {'category': '時間かかる', 'data_start': 6, 'hinmei_col': 4, 'hinban_col': 3, 'kosu_col': 12, 'nyuka_col': 1},
    '買い物': {'category': '買い物', 'data_start': 6, 'hinmei_col': 4, 'hinban_col': 3, 'kosu_col': 12, 'nyuka_col': 1},
    '副資材': {'category': '副資材', 'data_start': 6, 'hinmei_col': 5, 'hinban_col': 4, 'kosu_col': 13, 'nyuka_col': 1},
}


@pytest.fixture(scope='module')
def workbook():
    return openpyxl.load_workbook(EXCEL_PATH, data_only=True)


def _count_data_rows(ws, config):
    """Excelシートのデータ行数をカウント"""
    count = 0
    for row_idx in range(config['data_start'], ws.max_row + 1):
        hinmei = ws.cell(row=row_idx, column=config['hinmei_col']).value
        hinban = ws.cell(row=row_idx, column=config['hinban_col']).value
        if hinmei is None and hinban is None:
            continue
        count += 1
    return count


def _sum_kosu(ws, config):
    """Excelシートの検査工数(分)の合計を計算"""
    total = 0.0
    for row_idx in range(config['data_start'], ws.max_row + 1):
        hinmei = ws.cell(row=row_idx, column=config['hinmei_col']).value
        hinban = ws.cell(row=row_idx, column=config['hinban_col']).value
        if hinmei is None and hinban is None:
            continue
        val = ws.cell(row=row_idx, column=config['kosu_col']).value
        if val is not None:
            try:
                total += float(val)
            except (ValueError, TypeError):
                pass
    return total


def _read_csv_rows(csv_path):
    """CSVファイルを読み込み、ヘッダーとデータ行を返す"""
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = list(reader)
    return header, rows


# =============================================================
# 1. serial_to_date: 日付変換の単体テスト
# =============================================================
@pytest.mark.parametrize("serial,expected,should_succeed", [
    # 正常系: 既知のExcelシリアル値
    (45538, '2024-09-03', True),
    (45583, '2024-10-18', True),
    (45604, '2024-11-08', True),
    # 正常系: datetimeオブジェクト
    (datetime(2025, 1, 15), '2025-01-15', True),
    # 正常系: 40000以下の数値はそのまま文字列
    (100, '100', True),
    # 正常系: None → 空文字
    (None, '', True),
    # 正常系: 文字列はそのまま
    ('テスト', 'テスト', True),
    # 正常系: 浮動小数点のシリアル値（小数部は切り捨て）
    (45538.75, '2024-09-03', True),
    # 異常系: 空文字 → そのまま返す
    ('', '', True),
])
def test_serial_to_date(serial, expected, should_succeed):
    """Excelシリアル値を正しい日付文字列に変換する"""
    if should_succeed:
        assert serial_to_date(serial) == expected


# =============================================================
# 2. safe_filename: ファイル名変換の単体テスト
# =============================================================
@pytest.mark.parametrize("input_name,expected,should_succeed", [
    # 正常系: 通常の品名
    ('保護パッド', '保護パッド', True),
    ('Arm pipe Assy', 'Arm pipe Assy', True),
    # 正常系: 特殊文字を含む品名 → 置換
    ('Harness(FC-GPSヘリカル/PCM)', 'Harness(FC-GPSヘリカル_PCM)', True),
    ('test:file*name', 'test_file_name', True),
    ('path\\to"file', 'path_to_file', True),
    # 正常系: 改行を含む → 置換
    ('改行\nあり', '改行_あり', True),
    # 正常系: 空文字・None → '不明'
    ('', '不明', True),
    (None, '不明', True),
    # 正常系: 前後の空白 → strip
    ('  前後空白  ', '前後空白', True),
    # 正常系: 60文字超 → 切り詰め
    ('A' * 80, 'A' * 60, True),
])
def test_safe_filename(input_name, expected, should_succeed):
    """品名をファイルシステムで安全なファイル名に変換する"""
    if should_succeed:
        assert safe_filename(input_name) == expected


# =============================================================
# 3. 各シートの行数がExcelとCSVで一致するか
# =============================================================
@pytest.mark.parametrize("sheet_name,config,should_succeed",
    [(name, cfg, True) for name, cfg in SHEET_CONFIG.items()]
)
def test_row_count_matches(workbook, sheet_name, config, should_succeed):
    """Excelの各シートのデータ行数とCSVの行数が一致する"""
    ws = workbook[sheet_name]
    excel_count = _count_data_rows(ws, config)

    csv_path = os.path.join(RAW_DIR, f"{config['category']}_生データ.csv")
    _, csv_rows = _read_csv_rows(csv_path)

    if should_succeed:
        assert len(csv_rows) == excel_count, (
            f"{sheet_name}: Excel={excel_count}行, CSV={len(csv_rows)}行"
        )


# =============================================================
# 4. 検査工数(分)の合計がExcelとCSVで一致するか
# =============================================================
@pytest.mark.parametrize("sheet_name,config,should_succeed",
    [(name, cfg, True) for name, cfg in SHEET_CONFIG.items()]
)
def test_kosu_total_matches(workbook, sheet_name, config, should_succeed):
    """各シートの検査工数(分)合計がExcelとCSVで一致する"""
    ws = workbook[sheet_name]
    excel_total = _sum_kosu(ws, config)

    csv_path = os.path.join(RAW_DIR, f"{config['category']}_生データ.csv")
    header, csv_rows = _read_csv_rows(csv_path)

    kosu_idx = header.index('検査工数(分)')
    csv_total = 0.0
    for row in csv_rows:
        val = row[kosu_idx]
        if val:
            try:
                csv_total += float(val)
            except (ValueError, TypeError):
                pass

    if should_succeed:
        assert csv_total == pytest.approx(excel_total, abs=0.01), (
            f"{sheet_name}: Excel合計={excel_total}分, CSV合計={csv_total}分"
        )


# =============================================================
# 5. Excelにある品名がCSVに全て含まれるか
# =============================================================
@pytest.mark.parametrize("sheet_name,config,should_succeed",
    [(name, cfg, True) for name, cfg in SHEET_CONFIG.items()]
)
def test_no_missing_parts(workbook, sheet_name, config, should_succeed):
    """Excelにある品名がCSVから1つも欠落しない"""
    ws = workbook[sheet_name]
    excel_parts = set()
    for row_idx in range(config['data_start'], ws.max_row + 1):
        hinmei = ws.cell(row=row_idx, column=config['hinmei_col']).value
        hinban = ws.cell(row=row_idx, column=config['hinban_col']).value
        if hinmei is None and hinban is None:
            continue
        part = str(hinmei or hinban).strip()
        excel_parts.add(part)

    csv_path = os.path.join(RAW_DIR, f"{config['category']}_生データ.csv")
    header, csv_rows = _read_csv_rows(csv_path)

    hinmei_idx = header.index('品名') if '品名' in header else None
    hinban_idx = header.index('品番') if '品番' in header else None

    csv_parts = set()
    for row in csv_rows:
        part = ''
        if hinmei_idx is not None and row[hinmei_idx]:
            part = row[hinmei_idx].strip()
        elif hinban_idx is not None and row[hinban_idx]:
            part = row[hinban_idx].strip()
        if part:
            csv_parts.add(part)

    if should_succeed:
        missing = excel_parts - csv_parts
        assert len(missing) == 0, (
            f"{sheet_name}: CSVに欠落している品名: {missing}"
        )


# =============================================================
# 6. CSVの日付列がExcel実データと一致するか
# =============================================================
@pytest.mark.parametrize("sheet_name,config,should_succeed",
    [(name, cfg, True) for name, cfg in SHEET_CONFIG.items()]
)
def test_csv_dates_match_excel(workbook, sheet_name, config, should_succeed):
    """CSVに書き出された日付がExcelの元データと一致する"""
    ws = workbook[sheet_name]
    csv_path = os.path.join(RAW_DIR, f"{config['category']}_生データ.csv")
    header, csv_rows = _read_csv_rows(csv_path)

    nyuka_idx = header.index('入荷日')

    # Excelから日付を取得（データ行のみ、空行スキップ）
    excel_dates = []
    for row_idx in range(config['data_start'], ws.max_row + 1):
        hinmei = ws.cell(row=row_idx, column=config['hinmei_col']).value
        hinban = ws.cell(row=row_idx, column=config['hinban_col']).value
        if hinmei is None and hinban is None:
            continue
        val = ws.cell(row=row_idx, column=config['nyuka_col']).value
        excel_dates.append(serial_to_date(val))

    csv_dates = [row[nyuka_idx] for row in csv_rows]

    if should_succeed:
        assert csv_dates == excel_dates, (
            f"{sheet_name}: 日付が一致しない箇所がある"
        )


# =============================================================
# 7. CSVのカテゴリ列が正しいシート名に対応しているか
# =============================================================
@pytest.mark.parametrize("sheet_name,config,should_succeed",
    [(name, cfg, True) for name, cfg in SHEET_CONFIG.items()]
)
def test_category_column_correct(workbook, sheet_name, config, should_succeed):
    """raw CSVの全行のカテゴリ列が正しいシート由来になっている"""
    csv_path = os.path.join(RAW_DIR, f"{config['category']}_生データ.csv")
    header, csv_rows = _read_csv_rows(csv_path)

    cat_idx = header.index('カテゴリ')
    expected_category = config['category']

    if should_succeed:
        wrong_rows = [i for i, row in enumerate(csv_rows, 1)
                      if row[cat_idx] != expected_category]
        assert len(wrong_rows) == 0, (
            f"{sheet_name}: カテゴリが '{expected_category}' でない行: {wrong_rows[:5]}"
        )


# =============================================================
# 8. CSVヘッダーの存在チェック
# =============================================================
@pytest.mark.parametrize("csv_name,required_headers,should_succeed", [
    # 正常系: raw CSVに必須ヘッダーがあるか
    ('メカ_生データ.csv', ['カテゴリ', '品番', '品名', '検査工数(分)', '入荷日', '検査完了日'], True),
    ('エレキ_生データ.csv', ['カテゴリ', '品番', '品名', '検査工数(分)', '入荷日', '検査完了日'], True),
    ('Api_生データ.csv', ['カテゴリ', '品番', '品名', '検査工数(分)', '入荷日', '検査完了日'], True),
    # 異常系: 存在しないヘッダー
    ('メカ_生データ.csv', ['存在しない列'], False),
])
def test_raw_csv_headers(csv_name, required_headers, should_succeed):
    """raw CSVに必須ヘッダーが含まれる"""
    csv_path = os.path.join(RAW_DIR, csv_name)
    header, _ = _read_csv_rows(csv_path)
    if should_succeed:
        for h in required_headers:
            assert h in header, f"{csv_name}にヘッダー '{h}' がない"
    else:
        missing = [h for h in required_headers if h not in header]
        assert len(missing) > 0, f"{csv_name}: 存在しないはずのヘッダーが全て見つかった"


# =============================================================
# 9. by-part CSVの合計行数がrawと一致するか
# =============================================================
@pytest.mark.parametrize("should_succeed", [True])
def test_bypart_total_rows_match_raw(should_succeed):
    """品名別CSVの全行数合計がrawの全行数合計と一致する"""
    raw_total = 0
    for fname in os.listdir(RAW_DIR):
        if fname.endswith('.csv'):
            _, rows = _read_csv_rows(os.path.join(RAW_DIR, fname))
            raw_total += len(rows)

    part_total = 0
    for fname in os.listdir(PART_DIR):
        if fname.endswith('.csv'):
            _, rows = _read_csv_rows(os.path.join(PART_DIR, fname))
            part_total += len(rows)

    if should_succeed:
        assert part_total == raw_total, (
            f"raw合計={raw_total}行, by-part合計={part_total}行"
        )


# =============================================================
# 10. by-part CSVの工数合計がrawと一致するか
# =============================================================
@pytest.mark.parametrize("should_succeed", [True])
def test_bypart_kosu_total_matches_raw(should_succeed):
    """品名別CSVの工数合計がrawの工数合計と一致する"""
    def _sum_kosu_from_csvs(directory):
        total = 0.0
        for fname in os.listdir(directory):
            if not fname.endswith('.csv'):
                continue
            header, rows = _read_csv_rows(os.path.join(directory, fname))
            kosu_idx = header.index('検査工数(分)')
            for row in rows:
                val = row[kosu_idx]
                if val:
                    try:
                        total += float(val)
                    except (ValueError, TypeError):
                        pass
        return total

    raw_total = _sum_kosu_from_csvs(RAW_DIR)
    part_total = _sum_kosu_from_csvs(PART_DIR)

    if should_succeed:
        assert part_total == pytest.approx(raw_total, abs=0.01), (
            f"raw工数合計={raw_total}分, by-part工数合計={part_total}分"
        )


# =============================================================
# 11. summary CSVの集計値が正しいか
# =============================================================
@pytest.mark.parametrize("should_succeed", [True])
def test_summary_kosu_total_matches_raw(should_succeed):
    """summaryの合計工数がrawの全シートの工数合計と一致する"""
    raw_kosu_total = 0.0
    for fname in os.listdir(RAW_DIR):
        if not fname.endswith('.csv'):
            continue
        header, rows = _read_csv_rows(os.path.join(RAW_DIR, fname))
        kosu_idx = header.index('検査工数(分)')
        for row in rows:
            val = row[kosu_idx]
            if val:
                try:
                    raw_kosu_total += float(val)
                except (ValueError, TypeError):
                    pass

    summary_path = os.path.join(SUMMARY_DIR, '品名別_検査工数集計.csv')
    header, rows = _read_csv_rows(summary_path)
    kosu_idx = header.index('合計工数(分)')

    summary_kosu_total = 0.0
    for row in rows:
        if not row or not row[0] or row[0] == '【合計】':
            continue
        val = row[kosu_idx]
        if val:
            try:
                summary_kosu_total += float(val)
            except (ValueError, TypeError):
                pass

    if should_succeed:
        assert summary_kosu_total == pytest.approx(raw_kosu_total, abs=0.01), (
            f"raw工数合計={raw_kosu_total}分, summary工数合計={summary_kosu_total}分"
        )


@pytest.mark.parametrize("should_succeed", [True])
def test_summary_record_count_matches_raw(should_succeed):
    """summaryの検査回数合計がrawの全行数と一致する"""
    raw_total = 0
    for fname in os.listdir(RAW_DIR):
        if fname.endswith('.csv'):
            _, rows = _read_csv_rows(os.path.join(RAW_DIR, fname))
            raw_total += len(rows)

    summary_path = os.path.join(SUMMARY_DIR, '品名別_検査工数集計.csv')
    header, rows = _read_csv_rows(summary_path)
    count_idx = header.index('検査回数')

    summary_count = 0
    for row in rows:
        if not row or not row[0] or row[0] == '【合計】':
            continue
        val = row[count_idx]
        if val:
            try:
                summary_count += int(val)
            except (ValueError, TypeError):
                pass

    if should_succeed:
        assert summary_count == raw_total, (
            f"raw全行数={raw_total}, summary検査回数合計={summary_count}"
        )


# =============================================================
# 12. summaryの平均工数が正しく計算されているか
# =============================================================
@pytest.mark.parametrize("should_succeed", [True])
def test_summary_average_kosu_correct(should_succeed):
    """summaryの平均工数(分/回)が合計÷工数記録ありの件数と一致する"""
    summary_path = os.path.join(SUMMARY_DIR, '品名別_検査工数集計.csv')
    header, rows = _read_csv_rows(summary_path)

    kosu_idx = header.index('合計工数(分)')
    valid_idx = header.index('工数記録あり')
    avg_idx = header.index('平均工数(分/回)')

    errors = []
    for row in rows:
        if not row or not row[0] or row[0] == '【合計】':
            continue
        try:
            total = float(row[kosu_idx]) if row[kosu_idx] else 0
            valid = int(row[valid_idx]) if row[valid_idx] else 0
            avg = float(row[avg_idx]) if row[avg_idx] else 0
        except (ValueError, TypeError):
            continue

        if valid > 0:
            expected_avg = round(total / valid, 1)
            if abs(avg - expected_avg) > 0.01:
                errors.append(f"{row[0]}: 平均={avg}, 期待={expected_avg}")

    if should_succeed:
        assert len(errors) == 0, f"平均工数の計算ミス: {errors}"


# =============================================================
# 13. summaryが合計工数の降順でソートされているか
# =============================================================
@pytest.mark.parametrize("should_succeed", [True])
def test_summary_sorted_by_kosu_descending(should_succeed):
    """summaryの品名が合計工数(分)の降順で並んでいる"""
    summary_path = os.path.join(SUMMARY_DIR, '品名別_検査工数集計.csv')
    header, rows = _read_csv_rows(summary_path)

    kosu_idx = header.index('合計工数(分)')

    values = []
    for row in rows:
        if not row or not row[0] or row[0] == '【合計】':
            continue
        try:
            values.append(float(row[kosu_idx]) if row[kosu_idx] else 0)
        except (ValueError, TypeError):
            values.append(0)

    if should_succeed:
        assert values == sorted(values, reverse=True), (
            "summaryが合計工数の降順でソートされていない"
        )


# =============================================================
# 14. サンプルデータによるスポットチェック（空行スキップの独立検証）
# =============================================================
@pytest.mark.parametrize("category,expected_first_part,should_succeed", [
    # 正常系: 各シートの1行目の品名がExcelの6行目と一致
    ('メカ', '保護パッド', True),
    ('エレキ', 'Harness（Battery）', True),
    ('Api', 'Lower Plate（AC102）', True),
])
def test_first_row_spot_check(category, expected_first_part, should_succeed):
    """各シートCSVの1行目がExcelのデータ開始行と一致する"""
    csv_path = os.path.join(RAW_DIR, f"{category}_生データ.csv")
    header, csv_rows = _read_csv_rows(csv_path)
    hinmei_idx = header.index('品名')

    if should_succeed:
        assert csv_rows[0][hinmei_idx] == expected_first_part, (
            f"{category}: 1行目の品名が '{expected_first_part}' でなく '{csv_rows[0][hinmei_idx]}'"
        )
