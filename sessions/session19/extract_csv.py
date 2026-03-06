"""受入検査作業集計.xlsx から CSV を抽出するスクリプト

Step 1: 各シートの生データをCSV出力
Step 2: 品名ごとにCSV分割（時間データ付き）
Step 3: 品名ごとの時間集計CSVを出力
"""

import openpyxl
import csv
import os
import re
from datetime import datetime, timedelta
from collections import defaultdict

EXCEL_PATH = '/home/fuji0130/workspace/QA-FY26/docs/excel-orgn/受入検査作業集計.xlsx'
OUT_DIR = '/home/fuji0130/workspace/QA-FY26/sessions/session19/csv-output'

# Excelシリアル値を日付文字列に変換
def serial_to_date(val):
    if isinstance(val, (int, float)) and val > 40000:
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(val))).strftime('%Y-%m-%d')
        except:
            return str(val)
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val) if val is not None else ''

# 各シートのヘッダー行と列マッピング定義
# (シート名, ヘッダー行, 品番列, 品名列, 検査工数列, カテゴリ名)
SHEET_CONFIG = {
    '検品結果報告（メカ）': {
        'header_row': 5,
        'data_start': 6,
        'col_map': {
            '入荷日': 1, '検査完了日': 2, '発注番号': 3, '品番': 4, '品名': 5,
            'サプライヤ': 6, '入荷数量': 7, '不良数量': 8, '検査内容': 9,
            '検査資料': 10, '作業者': 11, '検査数量': 12, '検査工数(分)': 13,
            '人工': 14, '検査場所': 15, '検査治具': 16, '検査結果・備考': 17,
            '品質基準書修正要否と箇所': 18
        },
        'category': 'メカ'
    },
    '検品結果報告（エレキ）': {
        'header_row': 5,
        'data_start': 6,
        'col_map': {
            '入荷日': 1, '検査完了日': 2, '品番': 3, '品名': 4,
            'サプライヤ': 5, '入荷数量': 6, '不良数量': 7, '検査内容': 8,
            '検査資料': 9, '作業者': 10, '検査数量': 11, '検査工数(分)': 12,
            '人工': 13, '検査場所': 14, '検査治具': 15, '検査結果・備考': 16,
            '品質基準書修正要否と箇所': 17
        },
        'category': 'エレキ'
    },
    '検品結果報告（Api）': {
        'header_row': 5,
        'data_start': 6,
        'col_map': {
            '入荷日': 1, '検査完了日': 2, '品番': 3, '品名': 4,
            'サプライヤ': 5, '入荷数量': 6, '不良数量': 7, '検査内容': 8,
            '検査資料': 9, '作業者': 10, '検査数量': 11, '検査工数(分)': 12,
            '検査場所': 14, '検査治具': 15, '検査結果・備考': 16, 'その他': 17
        },
        'category': 'Api'
    },
    '受入に時間がかかるもの': {
        'header_row': 5,
        'data_start': 6,
        'col_map': {
            '入荷日': 1, '検査完了日': 2, '品番': 3, '品名': 4,
            'サプライヤ': 5, '入荷数量': 6, '不良数量': 7, '検査内容': 8,
            '検査資料': 9, '作業者': 10, '検査数量': 11, '検査工数(分)': 12,
            '検査場所': 14, '検査治具': 15, '検査結果・備考': 16
        },
        'category': '時間かかる'
    },
    '買い物': {
        'header_row': 5,
        'data_start': 6,
        'col_map': {
            '入荷日': 1, '検査完了日': 2, '品番': 3, '品名': 4,
            'サプライヤ': 5, '入荷数量': 6, '不良数量': 7, '検査内容': 8,
            '検査資料': 9, '作業者': 10, '検査数量': 11, '検査工数(分)': 12,
            '人工': 13, '検査場所': 14, '検査治具': 15, '検査結果・備考': 16
        },
        'category': '買い物'
    },
    '副資材': {
        'header_row': 5,
        'data_start': 6,
        'col_map': {
            '入荷日': 1, '検査完了日': 2, '発注番号': 3, '品番': 4, '品名': 5,
            'サプライヤ': 6, '入荷数量': 7, '不良数量': 8, '検査内容': 9,
            '検査資料': 10, '作業者': 11, '検査数量': 12, '検査工数(分)': 13,
            '人工': 14, '検査場所': 15, '検査治具': 16, '検査結果・備考': 17
        },
        'category': '副資材'
    },
}


def safe_filename(name):
    """ファイル名に使えない文字を置換"""
    if not name:
        return '不明'
    name = str(name).strip()
    # ファイルシステムで問題になる文字を置換
    name = re.sub(r'[\\/:*?"<>|]', '_', name)
    name = name.replace('\n', '_').replace('\r', '')
    # 長すぎる名前を切り詰め
    if len(name) > 60:
        name = name[:60]
    return name


def get_cell_value(ws, row, col):
    """セルの値を取得（None対応）"""
    val = ws.cell(row=row, column=col).value
    return val


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ===== Step 1: 各シートの生データをCSV出力 =====
    print("=" * 60)
    print("Step 1: 生データCSV出力")
    print("=" * 60)

    raw_dir = os.path.join(OUT_DIR, 'raw')
    all_records = []  # Step 2, 3 用に全レコードを蓄積

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [SKIP] シート '{sheet_name}' が見つかりません")
            continue

        ws = wb[sheet_name]
        category = config['category']
        col_map = config['col_map']
        data_start = config['data_start']

        # ヘッダー名のリスト（列順にソート）
        headers_sorted = sorted(col_map.items(), key=lambda x: x[1])
        header_names = [h[0] for h in headers_sorted]
        header_cols = [h[1] for h in headers_sorted]

        # CSV書き出し
        csv_path = os.path.join(raw_dir, f'{category}_生データ.csv')
        row_count = 0

        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # ヘッダー + カテゴリ列
            writer.writerow(['カテゴリ'] + header_names)

            for row_idx in range(data_start, ws.max_row + 1):
                # 空行スキップ（品名 or 品番が空なら）
                hinmei_col = col_map.get('品名', col_map.get('品番'))
                if get_cell_value(ws, row_idx, hinmei_col) is None:
                    # 品番も品名も無ければスキップ
                    hinban_col = col_map.get('品番')
                    if hinban_col and get_cell_value(ws, row_idx, hinban_col) is None:
                        continue

                row_data = [category]
                for col_idx in header_cols:
                    val = get_cell_value(ws, row_idx, col_idx)
                    # 日付列の変換
                    if col_idx in [col_map.get('入荷日', -1), col_map.get('検査完了日', -1)]:
                        row_data.append(serial_to_date(val))
                    else:
                        row_data.append(str(val) if val is not None else '')
                writer.writerow(row_data)
                row_count += 1

                # 全レコード用に辞書化
                record = {'カテゴリ': category}
                for name, col_idx in col_map.items():
                    val = get_cell_value(ws, row_idx, col_idx)
                    if name in ('入荷日', '検査完了日'):
                        record[name] = serial_to_date(val)
                    else:
                        record[name] = val
                all_records.append(record)

        print(f"  [{category}] {row_count}件 → {csv_path}")

    # ===== Step 2: 品名ごとにCSV分割 =====
    print()
    print("=" * 60)
    print("Step 2: 品名ごとにCSV分割（時間データ付き）")
    print("=" * 60)

    part_dir = os.path.join(OUT_DIR, 'by-part')

    # 品名でグルーピング
    by_part = defaultdict(list)
    for rec in all_records:
        part_name = rec.get('品名', '') or rec.get('品番', '') or '不明'
        part_name = str(part_name).strip()
        by_part[part_name].append(rec)

    # 出力ヘッダー（時間分析に必要な列のみ）
    time_headers = ['カテゴリ', '入荷日', '検査完了日', '品番', '品名', 'サプライヤ',
                    '入荷数量', '検査数量', '検査工数(分)', '作業者', '検査内容', '検査結果・備考']

    for part_name, records in sorted(by_part.items()):
        fname = safe_filename(part_name)
        csv_path = os.path.join(part_dir, f'{fname}.csv')

        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(time_headers)
            for rec in records:
                row = [str(rec.get(h, '') or '') for h in time_headers]
                writer.writerow(row)

        print(f"  [{part_name}] {len(records)}件 → {fname}.csv")

    # ===== Step 3: 品名ごとの時間集計CSV =====
    print()
    print("=" * 60)
    print("Step 3: 品名ごとの時間集計CSV")
    print("=" * 60)

    summary_dir = os.path.join(OUT_DIR, 'summary')
    summary_path = os.path.join(summary_dir, '品名別_検査工数集計.csv')

    # 集計
    summary_data = {}
    for part_name, records in sorted(by_part.items()):
        total_minutes = 0
        valid_count = 0
        invalid_values = []
        categories = set()

        for rec in records:
            categories.add(rec.get('カテゴリ', ''))
            raw_val = rec.get('検査工数(分)')
            if raw_val is None or raw_val == '':
                continue
            try:
                minutes = float(raw_val)
                total_minutes += minutes
                valid_count += 1
            except (ValueError, TypeError):
                invalid_values.append(str(raw_val))

        summary_data[part_name] = {
            'カテゴリ': '/'.join(sorted(categories)),
            '検査回数': len(records),
            '工数記録あり': valid_count,
            '合計工数(分)': total_minutes,
            '合計工数(時間)': round(total_minutes / 60, 2) if total_minutes > 0 else 0,
            '平均工数(分)': round(total_minutes / valid_count, 1) if valid_count > 0 else 0,
            '数値でない値': ', '.join(invalid_values) if invalid_values else ''
        }

    # 合計工数降順でソート
    sorted_summary = sorted(summary_data.items(), key=lambda x: x[1]['合計工数(分)'], reverse=True)

    with open(summary_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['品名', 'カテゴリ', '検査回数', '工数記録あり',
                         '合計工数(分)', '合計工数(時間)', '平均工数(分/回)', '数値でない値'])

        grand_total = 0
        for part_name, stats in sorted_summary:
            writer.writerow([
                part_name,
                stats['カテゴリ'],
                stats['検査回数'],
                stats['工数記録あり'],
                stats['合計工数(分)'],
                stats['合計工数(時間)'],
                stats['平均工数(分)'],
                stats['数値でない値']
            ])
            grand_total += stats['合計工数(分)']

        # 合計行
        writer.writerow([])
        writer.writerow(['【合計】', '', sum(s['検査回数'] for _, s in sorted_summary),
                         sum(s['工数記録あり'] for _, s in sorted_summary),
                         grand_total, round(grand_total / 60, 2), '', ''])

    print(f"  集計結果 → {summary_path}")
    print(f"  品名数: {len(summary_data)}")
    print(f"  全レコード数: {len(all_records)}")
    print(f"  合計工数: {grand_total}分 ({round(grand_total/60, 1)}時間)")

    # トップ10表示
    print()
    print("  --- 工数トップ10 ---")
    for i, (name, stats) in enumerate(sorted_summary[:10], 1):
        print(f"  {i:2d}. {name}: {stats['合計工数(分)']}分 ({stats['合計工数(時間)']}h) × {stats['検査回数']}回")


if __name__ == '__main__':
    main()
